"""
Multi-format Attachment Handler — ingests files of various types and converts
them to text chunks for the source store.

Supported formats:
  - PDF (.pdf) — full text extraction via pypdf
  - Markdown (.md) — direct text read
  - Jupyter Notebook (.ipynb) — extract markdown cells + code comments
  - CSV (.csv) — generate text summary of dataset
  - Excel (.xlsx, .xls) — extract sheet data as text
  - Image (.jpg, .jpeg, .png) — OCR via qwen-vl (optional)
  - Audio (.mp3, .wav) — transcription via Whisper (optional)
  - Video (.mp4) — audio extraction + Whisper transcription (optional)
  - DOCX (.docx) — text extraction via python-docx
"""
import logging
from pathlib import Path
from enum import Enum
from typing import Optional

logger = logging.getLogger(__name__)


class AttachmentType(str, Enum):
    PDF = "pdf"
    MARKDOWN = "markdown"
    NOTEBOOK = "notebook"
    CSV = "csv"
    EXCEL = "excel"
    IMAGE = "image"
    AUDIO = "audio"
    VIDEO = "video"
    DOCX = "docx"
    UNKNOWN = "unknown"


# MIME type → AttachmentType mapping
MIME_MAP = {
    "application/pdf": AttachmentType.PDF,
    "text/markdown": AttachmentType.MARKDOWN,
    "text/x-markdown": AttachmentType.MARKDOWN,
    "application/json": AttachmentType.NOTEBOOK,
    "text/csv": AttachmentType.CSV,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": AttachmentType.EXCEL,
    "application/vnd.ms-excel": AttachmentType.EXCEL,
    "image/jpeg": AttachmentType.IMAGE,
    "image/png": AttachmentType.IMAGE,
    "image/webp": AttachmentType.IMAGE,
    "audio/mpeg": AttachmentType.AUDIO,
    "audio/wav": AttachmentType.AUDIO,
    "audio/ogg": AttachmentType.AUDIO,
    "video/mp4": AttachmentType.VIDEO,
    "video/webm": AttachmentType.VIDEO,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": AttachmentType.DOCX,
}

# Size limits per type (in bytes)
SIZE_LIMITS = {
    AttachmentType.PDF: 50 * 1024 * 1024,       # 50MB
    AttachmentType.MARKDOWN: 5 * 1024 * 1024,   # 5MB
    AttachmentType.NOTEBOOK: 20 * 1024 * 1024,  # 20MB
    AttachmentType.CSV: 10 * 1024 * 1024,       # 10MB
    AttachmentType.EXCEL: 20 * 1024 * 1024,     # 20MB
    AttachmentType.IMAGE: 10 * 1024 * 1024,     # 10MB
    AttachmentType.AUDIO: 200 * 1024 * 1024,    # 200MB
    AttachmentType.VIDEO: 500 * 1024 * 1024,    # 500MB
    AttachmentType.DOCX: 20 * 1024 * 1024,      # 20MB
}


class AttachmentHandler:
    """
    Processes uploaded files of various types and returns extracted text
    suitable for chunking and embedding.
    """

    def classify(self, mime_type: str, filename: str = "") -> AttachmentType:
        """Classify a file by MIME type, falling back to extension."""
        attachment_type = MIME_MAP.get(mime_type, AttachmentType.UNKNOWN)
        if attachment_type == AttachmentType.UNKNOWN and filename:
            ext = Path(filename).suffix.lower()
            ext_map = {
                ".pdf": AttachmentType.PDF,
                ".md": AttachmentType.MARKDOWN,
                ".markdown": AttachmentType.MARKDOWN,
                ".ipynb": AttachmentType.NOTEBOOK,
                ".csv": AttachmentType.CSV,
                ".xlsx": AttachmentType.EXCEL,
                ".xls": AttachmentType.EXCEL,
                ".jpg": AttachmentType.IMAGE,
                ".jpeg": AttachmentType.IMAGE,
                ".png": AttachmentType.IMAGE,
                ".webp": AttachmentType.IMAGE,
                ".mp3": AttachmentType.AUDIO,
                ".wav": AttachmentType.AUDIO,
                ".ogg": AttachmentType.AUDIO,
                ".mp4": AttachmentType.VIDEO,
                ".webm": AttachmentType.VIDEO,
                ".docx": AttachmentType.DOCX,
            }
            attachment_type = ext_map.get(ext, AttachmentType.UNKNOWN)
        return attachment_type

    def get_size_limit(self, attachment_type: AttachmentType) -> int:
        """Get the maximum allowed file size for a given type."""
        return SIZE_LIMITS.get(attachment_type, 10 * 1024 * 1024)  # default 10MB

    async def process(self, file_path: Path, mime_type: str,
                      filename: str, manual_metadata: dict = None) -> dict:
        """
        Process an uploaded file and return extracted text + metadata.

        Returns:
            dict with keys: title, text, source_type, metadata
        """
        attachment_type = self.classify(mime_type, filename)
        meta = manual_metadata or {}

        handlers = {
            AttachmentType.PDF: self._handle_pdf,
            AttachmentType.MARKDOWN: self._handle_markdown,
            AttachmentType.NOTEBOOK: self._handle_notebook,
            AttachmentType.CSV: self._handle_csv,
            AttachmentType.EXCEL: self._handle_excel,
            AttachmentType.IMAGE: self._handle_image,
            AttachmentType.AUDIO: self._handle_audio,
            AttachmentType.VIDEO: self._handle_video,
            AttachmentType.DOCX: self._handle_docx,
        }

        handler = handlers.get(attachment_type)
        if not handler:
            return {
                "error": f"Unsupported attachment type: {mime_type}",
                "source_type": "unknown",
            }

        try:
            return await handler(file_path, filename, meta)
        except Exception as e:
            logger.exception("Attachment processing failed for %s: %s", filename, e)
            return {
                "error": str(e),
                "title": meta.get("title", filename),
                "text": "",
                "source_type": attachment_type.value,
                "metadata": meta,
            }

    async def _handle_pdf(self, path: Path, filename: str, meta: dict) -> dict:
        """Extract text from PDF using pypdf."""
        from services.ingestion.pdf_extractor import PDFExtractor
        extractor = PDFExtractor()
        result = await extractor.extract(path)
        return {
            "title": meta.get("title", result.get("title", filename)),
            "text": result.get("text", ""),
            "source_type": "pdf",
            "metadata": {**meta, "authors": result.get("authors"), "year": result.get("year")},
        }

    async def _handle_markdown(self, path: Path, filename: str, meta: dict) -> dict:
        """Read markdown file directly."""
        text = path.read_text(encoding="utf-8")
        return {
            "title": meta.get("title", filename.replace(".md", "").replace(".markdown", "")),
            "text": text,
            "source_type": "markdown",
            "metadata": meta,
        }

    async def _handle_notebook(self, path: Path, filename: str, meta: dict) -> dict:
        """Extract text from Jupyter notebook: markdown cells + code comments."""
        import nbformat
        nb = nbformat.read(str(path), as_version=4)
        parts = []
        for cell in nb.cells:
            if cell.cell_type == "markdown":
                parts.append(cell.source)
            elif cell.cell_type == "code":
                # Extract docstrings and comments
                for line in cell.source.split("\n"):
                    stripped = line.strip()
                    if stripped.startswith("#"):
                        parts.append(stripped.lstrip("# "))
                    elif '"""' in stripped or "'''" in stripped:
                        parts.append(stripped)
        text = "\n\n".join(parts)
        return {
            "title": meta.get("title", filename.replace(".ipynb", "")),
            "text": text,
            "source_type": "notebook",
            "metadata": meta,
        }

    async def _handle_csv(self, path: Path, filename: str, meta: dict) -> dict:
        """Generate a text summary of a CSV dataset."""
        import csv
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        if not rows:
            return {
                "title": meta.get("title", filename),
                "text": f"Empty CSV: {filename}",
                "source_type": "csv",
                "metadata": meta,
            }

        summary = (
            f"Dataset: {filename}\n"
            f"Columns: {', '.join(rows[0].keys())}\n"
            f"Rows: {len(rows)}\n"
            f"Sample (first 5 rows):\n"
        )
        for row in rows[:5]:
            summary += str(dict(row)) + "\n"

        return {
            "title": meta.get("title", filename),
            "text": summary,
            "source_type": "csv",
            "metadata": meta,
        }

    async def _handle_excel(self, path: Path, filename: str, meta: dict) -> dict:
        """Extract text from Excel files."""
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"Sheet: {sheet_name}")
            for row in list(ws.iter_rows(values_only=True))[:30]:
                parts.append(str(row))
        return {
            "title": meta.get("title", filename),
            "text": "\n".join(parts),
            "source_type": "excel",
            "metadata": meta,
        }

    async def _handle_image(self, path: Path, filename: str, meta: dict) -> dict:
        """
        Extract text from image using qwen-vl.
        Falls back gracefully if qwen-vl is unavailable.
        """
        try:
            import base64
            from services.intelligence.qwen_client import QwenClient

            with open(path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()

            qwen = QwenClient()
            result = await qwen.complete(
                model="qwen-vl-max",
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
                        },
                        {
                            "type": "text",
                            "text": (
                                "Extract all text and mathematical equations from this image. "
                                "Return LaTeX for equations. Return plain text for prose. "
                                "Label sections. Do not interpret or summarise."
                            ),
                        },
                    ],
                }],
            )
            return {
                "title": meta.get("title", filename),
                "text": result,
                "source_type": "image_ocr",
                "metadata": meta,
            }
        except ImportError:
            logger.warning("qwen-vl not available for image OCR on %s", filename)
            return {
                "title": meta.get("title", filename),
                "text": f"[Image file: {filename}. OCR not available.]",
                "source_type": "image_ocr",
                "metadata": meta,
            }
        except Exception as e:
            logger.warning("Image OCR failed for %s: %s", filename, e)
            return {
                "title": meta.get("title", filename),
                "text": f"[Image file: {filename}. OCR failed: {e}]",
                "source_type": "image_ocr",
                "metadata": meta,
            }

    async def _handle_audio(self, path: Path, filename: str, meta: dict) -> dict:
        """
        Transcribe audio using Whisper.
        Falls back gracefully if whisper is not installed.
        """
        try:
            import whisper
            model = whisper.load_model("base")
            result = model.transcribe(str(path))
            return {
                "title": meta.get("title", filename),
                "text": result["text"],
                "source_type": "audio_transcript",
                "metadata": {**meta, "language": result.get("language", "en")},
            }
        except ImportError:
            logger.warning("whisper not available for audio transcription on %s", filename)
            return {
                "title": meta.get("title", filename),
                "text": f"[Audio file: {filename}. Transcription not available.]",
                "source_type": "audio_transcript",
                "metadata": meta,
            }
        except Exception as e:
            logger.warning("Audio transcription failed for %s: %s", filename, e)
            return {
                "title": meta.get("title", filename),
                "text": f"[Audio file: {filename}. Transcription failed: {e}]",
                "source_type": "audio_transcript",
                "metadata": meta,
            }

    async def _handle_video(self, path: Path, filename: str, meta: dict) -> dict:
        """
        Extract audio from video with ffmpeg, then transcribe with Whisper.
        Falls back gracefully if either tool is unavailable.
        """
        import subprocess
        import tempfile

        try:
            with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as tmp:
                tmp_path = Path(tmp.name)

            subprocess.run(
                ["ffmpeg", "-i", str(path), "-vn", "-acodec", "pcm_s16le",
                 "-ar", "16000", "-ac", "1", str(tmp_path), "-y"],
                capture_output=True,
                check=True,
                timeout=300,
            )

            result = await self._handle_audio(tmp_path, filename, meta)
            tmp_path.unlink(missing_ok=True)
            result["source_type"] = "video_transcript"
            return result
        except FileNotFoundError:
            logger.warning("ffmpeg not available for video processing on %s", filename)
            return {
                "title": meta.get("title", filename),
                "text": f"[Video file: {filename}. ffmpeg not available.]",
                "source_type": "video_transcript",
                "metadata": meta,
            }
        except subprocess.TimeoutExpired:
            logger.warning("ffmpeg timed out processing %s", filename)
            return {
                "title": meta.get("title", filename),
                "text": f"[Video file: {filename}. Processing timed out.]",
                "source_type": "video_transcript",
                "metadata": meta,
            }
        except Exception as e:
            logger.warning("Video processing failed for %s: %s", filename, e)
            return {
                "title": meta.get("title", filename),
                "text": f"[Video file: {filename}. Processing failed: {e}]",
                "source_type": "video_transcript",
                "metadata": meta,
            }

    async def _handle_docx(self, path: Path, filename: str, meta: dict) -> dict:
        """Extract text from DOCX files."""
        from docx import Document
        doc = Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return {
            "title": meta.get("title", filename),
            "text": text,
            "source_type": "docx",
            "metadata": meta,
        }